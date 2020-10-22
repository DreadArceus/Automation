#Bad PDF, resulting in random errors in converted xl file
#Dropped the idea for that reason and because finding the add button will be hard
#Should get better at selenium someday
from openpyxl import load_workbook
import time
wb = load_workbook(filename = 'jossa.xlsx')
ws = wb['Table 1']
for i in range(410):
    print(ws['B' + f'{i + 54}'].value + ws['D' + f'{i + 54}'].value)
    time.sleep(1)