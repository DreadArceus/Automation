from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook(filename = str(input()))
sheets = wb.sheetnames
months = [[x, 0] for x in range(68, 77)]
wr = Workbook()
for s in sheets:
    ws = wb[s]
    sr, sc = 3, 2
    result = wr.create_sheet(title = 'DefaulterList')
    currentClass = ws['C' + f'{sr}'].value
    for m in months:
        result[chr(m[0] - 2) + '1'] = ws[chr(m[0]) + f'2'].value + ' Till Date'
    while(ws['C' + f'{sr}'].value):
        if ws['C' + f'{sr}'].value != currentClass:
            result['A' + f'{sc}'] = currentClass
            for m in months:
                result[chr(m[0] - 2) + f'{sc}'] = m[1]
            months = [[x, 0] for x in range(68, 77)]
            currentClass = ws['C' + f'{sr}'].value
            sc += 1
        stdM = -1
        for i in range(len(months)):
            if ws[chr(months[i][0]) + f'{sr}'].value:
                stdM = i
                break
        if stdM != -1:
            months[stdM][1] += 1
        sr += 1
    result['A' + f'{sc}'] = currentClass
    for m in months:
        result[chr(m[0] - 2) + f'{sc}'] = m[1]
wr.save(filename = 'wR.xlsx')