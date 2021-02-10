from openpyxl import load_workbook
from openpyxl import Workbook
def convertToColumn(day):
    if(day < 27):
        return chr(64 + day)
    elif(day < 53):
        return 'A' + chr(38 + day)
    else:
        return 'B' + chr(12 + day)
def checkAbsence(day, sr):
    if(day != 0):
        day *= 2
    c1 = convertToColumn(day + 6)
    c2 = convertToColumn(day + 7)
    return ws[c1 + f'{sr}'].value == 'A' and ws[c2 + f'{sr}'].value == 'A'

months = ['JAN 21']
wb = load_workbook(filename = str(input()))
wr = Workbook()
for s in months:
    ws = wb[s]
    result = wr.create_sheet(title = "Result " + s)
    for day in range(31):
        conDay = convertToColumn(day + 1)
        result[conDay + '1'] = f'{day+1}'
        sr, sn = 4, 1
        while(ws['B' + f'{sr}'].value):
            if(checkAbsence(day, sr)):
                result[conDay + f'{sn+1}'] = ws['B' + f'{sr}'].value
                sn += 1
            sr += 1
wr.save(filename = 'wR.xlsx')