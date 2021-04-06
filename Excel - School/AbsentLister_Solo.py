from openpyxl import load_workbook
from openpyxl import Workbook

def convertToColumn(day):
    if(day < 27):
        return chr(64 + day)
    elif(day < 53):
        return 'A' + chr(38 + day)
    else:
        return 'B' + chr(12 + day)
def checkAbsence(day, sr):
    c = convertToColumn(day + 5)
    return ws[c + f'{sr}'].value == 'A'

wb = load_workbook(filename = str(input()))
months = ['FEB 2021']
wr = Workbook()
for s in months:
    ws = wb[s]
    result = wr.create_sheet(title = "Result " + s)
    for day in range(31):
        conDay = convertToColumn(day + 1)
        result[conDay + '1'] = f'{day+1}' + ' ' + s[0] + s[1] + s[2]
        sr, sn = 4, 1
        while(ws['B' + f'{sr}'].value):
            if(checkAbsence(day, sr)):
                result[conDay + f'{sn+1}'] = ws['B' + f'{sr}'].value
                sn += 1
            sr += 1
wr.save(filename = 'R.xlsx')