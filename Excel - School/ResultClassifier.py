from openpyxl import load_workbook
from openpyxl import Workbook
from os import walk

opt = int(input("Enter 1 for 80 scaling, 2 for 50 scaling: "))
folder_name = str(input("Enter folder name: "))
for (_, _, filenames) in walk(folder_name):
    for name in filenames:
        if ".xlsx" not in name:
            continue
        wb = load_workbook(filename=folder_name+"/"+name)
        ws = wb.worksheets[0].title
        ws = wb[ws]
        wr = Workbook()
        result = wr.create_sheet(title="Result")

        subjects = {}
        starting_column = "D"
        while ws[starting_column + '4'].value:
            subjects[ws[starting_column + '4'].value] = {"column": starting_column, "distribution": [0 for _ in range(8 if opt == 1 else 5)],
                                                         "passing": 0, "failing": 0}
            starting_column = chr(ord(starting_column) + 1)

        for i in range(8 if opt == 1 else 5):
            result[f'{chr(ord("A") + i + 1)}' + '1'] = f'{10 * i} to {10 * (i + 1)}'
        result[f'{chr(ord("A") + (8 if opt == 1 else 5) + 1)}' + "1"] = "Pass"
        result[f'{chr(ord("A") + (8 if opt == 1 else 5) + 2)}' + "1"] = "Fail"

        for key, value in subjects.items():
            sr = 9
            try:
                max_marks = int(str(ws[value["column"] + f'{sr - 1}'].value)[1:])
            except ValueError:
                continue
            passing_marks = 0
            scale = 1
            if opt == 1:
                passing_marks = 28
                scale = 80 / max_marks
            elif opt == 2:
                passing_marks = 17.5
                scale = 50 / max_marks
            while ws[value["column"] + f'{sr}'].value or ws[value["column"] + f'{sr}'].value == 0:
                try:
                    marks = float(ws[value["column"] + f'{sr}'].value) * scale
                except ValueError:
                    if ws[value["column"] + f'{sr}'].value == '---':
                        sr += 2
                        continue
                    marks = 0
                if marks >= passing_marks:
                    value["passing"] += 1
                else:
                    value["failing"] += 1
                value["distribution"][min(int(marks / 10), int(max_marks * scale / 10 - 1))] += 1
                sr += 2

            result["A" + f'{ord(value["column"]) - ord("A")}'] = key
            for i in range(len(value["distribution"])):
                result[f'{chr(ord("A") + 1 + i)}' + f'{ord(value["column"]) - ord("A")}'] = value["distribution"][i]
            result[f'{chr(ord("A") + (8 if opt == 1 else 5) + 1)}' + f'{ord(value["column"]) - ord("A")}'] = value["passing"]
            result[f'{chr(ord("A") + (8 if opt == 1 else 5) + 2)}' + f'{ord(value["column"]) - ord("A")}'] = value["failing"]

        wr.save(filename=f'output_{name}')
